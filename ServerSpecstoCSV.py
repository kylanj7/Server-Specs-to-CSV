import openpyxl
import subprocess
import datetime
import os
from openpyxl.styles import PatternFill, Border, Side, Alignment

def get_ram_info():
    """Get detailed RAM information"""
    try:
        # Use free command for total RAM
        free_cmd = "free -g"  # Get memory in GB directly
        free_output = subprocess.check_output(free_cmd.split(), universal_newlines=True)
        total_gb = int(free_output.splitlines()[1].split()[1])
        
        # Get Memory Device info for RAM type
        device_cmd = "sudo dmidecode -t 17"  # Type 17 is Memory Device
        device_info = subprocess.check_output(device_cmd.split(), universal_newlines=True)
        
        # Initialize variables
        ram_type = None
        
        # Parse RAM type from the first populated memory module
        current_device = False
        for line in device_info.split('\n'):
            if 'Memory Device' in line:
                current_device = True
            elif current_device:
                if 'Type:' in line and not ram_type:
                    ram_type = line.split(':')[1].strip()
                    if ram_type and ram_type != "Unknown":
                        break
        
        return ram_type, f"{total_gb}GB"
    except Exception as e:
        print(f"Error getting RAM info: {e}")
        return "Unknown", "Unknown"

def get_pcie_devices():
    """Get PCIe device information"""
    try:
        # Use lspci command to get detailed device information
        pcie_cmd = "lspci -vmm"
        pcie_info = subprocess.check_output(pcie_cmd.split(), universal_newlines=True)
        
        # Parse the output to get device information
        devices = []
        current_device = {}
        
        for line in pcie_info.split('\n'):
            if line.strip():
                if ':' in line:
                    key, value = line.split(':', 1)
                    key = key.strip()
                    value = value.strip()
                    
                    if key == 'Device':
                        current_device['model'] = value
                    elif key == 'Class':
                        current_device['type'] = value
            else:
                if current_device:
                    # Only add devices we're interested in (excluding bridges, etc.)
                    if 'type' in current_device and any(x in current_device['type'].lower() 
                        for x in ['storage', 'ethernet', 'fibre', 'raid', 'sas']):
                        devices.append(current_device)
                current_device = {}
        
        # Format device strings
        device_strings = []
        for device in devices:
            if 'model' in device and 'type' in device:
                device_strings.append(f"{device['model']} ({device['type']})")
        
        return device_strings
    except Exception as e:
        print(f"Error getting PCIe info: {e}")
        return []

def get_cpu_info():
    """Get CPU information using dmidecode"""
    try:
        cpu_cmd = "sudo dmidecode -t processor"
        cpu_info = subprocess.check_output(cpu_cmd.split(), universal_newlines=True)
        
        # Parse CPU information
        for line in cpu_info.split('\n'):
            if 'Version:' in line:
                return line.split(':')[1].strip()
        return "Unknown CPU"
    except Exception as e:
        print(f"Error getting CPU info: {e}")
        return "Unknown CPU"

def get_server_info():
    """Gather server information using dmidecode"""
    try:
        # Initialize all variables at the start
        manufacturer = None
        model = None
        serial = None
        product_number = None
        
        # Get System Information
        system_cmd = "sudo dmidecode -t system"
        board_cmd = "sudo dmidecode -t baseboard"
        
        try:
            system_info = subprocess.check_output(system_cmd.split(), universal_newlines=True)
            board_info = subprocess.check_output(board_cmd.split(), universal_newlines=True)
        except subprocess.CalledProcessError as e:
            print(f"Error executing dmidecode: {e}")
            return None
        
        # Parse system info
        for line in system_info.split('\n'):
            if 'Manufacturer:' in line:
                manufacturer = line.split(':')[1].strip()
            elif 'Product Name:' in line:
                model = line.split(':')[1].strip()
            elif 'Serial Number:' in line:
                serial = line.split(':')[1].strip()
        
        # Get P/N from baseboard info
        for line in board_info.split('\n'):
            if 'Product Name:' in line:
                product_number = line.split(':')[1].strip()
            
        # Get CPU, RAM, and PCIe info
        cpu_info = get_cpu_info()
        ram_type, installed_ram = get_ram_info()
        pcie_devices = get_pcie_devices()
        
        # Print found information for verification
        print("\nFound System Information:")
        print(f"Model: {model or 'Not Found'}")
        print(f"P/N: {product_number or 'Not Found'}")
        print(f"Serial: {serial or 'Not Found'}")
        print(f"CPU: {cpu_info}")
        print(f"RAM Type: {ram_type}")
        print(f"Installed RAM: {installed_ram}")
        print("\nPCIe Devices:")
        for device in pcie_devices:
            print(f"- {device}")
        
        verify = input("\nIs this information correct? (y/n): ").lower()
        if verify != 'y':
            print("\nPlease enter the correct information:")
            if input("Enter information manually? (y/n): ").lower() == 'y':
                model = input("Enter Model: ") or model
                product_number = input("Enter P/N: ") or product_number
                serial = input("Enter Serial Number: ") or serial
                cpu_info = input("Enter CPU info: ") or cpu_info
                ram_type = input("Enter RAM Type: ") or ram_type
                installed_ram = input("Enter Installed RAM: ") or installed_ram
                if input("Edit PCIe devices? (y/n): ").lower() == 'y':
                    pcie_devices = []
                    while True:
                        device = input("Enter PCIe device (or press Enter to finish): ")
                        if not device:
                            break
                        pcie_devices.append(device)
        
        return {
            'model': model or 'Unknown',
            'product_number': product_number or 'Unknown',
            'serial_number': serial or 'Unknown',
            'processor': cpu_info,
            'ram_type': ram_type,
            'installed_ram': installed_ram,
            'pcie_devices': pcie_devices
        }
        
    except Exception as e:
        print(f"Error gathering server information: {e}")
        return None

def validate_grade(grade_str, grade_type):
    """Validate grade format"""
    if grade_type == 'cosmetic':
        if not grade_str.startswith('C'):
            return False
    elif grade_type == 'functional':
        if not grade_str.startswith('F'):
            return False
    
    # Check if format is correct (C1-9 or F1-9)
    if len(grade_str) != 2:
        return False
    
    # Check if number is 1-9
    if not grade_str[1].isdigit() or int(grade_str[1]) < 1 or int(grade_str[1]) > 9:
        return False
        
    return True

def get_user_input():
    """Get required user input for QC process"""
    print("\nServer QC Data Collection")
    print("=" * 30)
    
    # Get cosmetic grade
    while True:
        cosmetic_grade = input("\nEnter Cosmetic Grade (C1-C9): ").upper()
        if validate_grade(cosmetic_grade, 'cosmetic'):
            break
        print("Invalid grade. Please enter a grade between C1 and C9.")
    
    # Get functional grade
    while True:
        functional_grade = input("Enter Functional Grade (F1-F9): ").upper()
        if validate_grade(functional_grade, 'functional'):
            break
        print("Invalid grade. Please enter a grade between F1 and F9.")
    
    # Get QC technician name
    qced_by = input("Enter QC Technician Name: ")
    
    # Get test date manually and format it with slashes
    while True:
        test_date = input("Enter Test Date (MMDDYY): ")
        if len(test_date) == 6 and test_date.isdigit():
            # Format the date with slashes (MM/DD/YY)
            formatted_date = f"{test_date[:2]}/{test_date[2:4]}/{test_date[4:]}"
            break
        print("Invalid date format. Please use MMDDYY format (e.g., 120924)")
    
    return {
        'cosmetic_grade': cosmetic_grade,
        'functional_grade': functional_grade,
        'qced_by': qced_by,
        'test_date': formatted_date
    }

def create_qc_spreadsheet(data):
    """Create and populate QC spreadsheet"""
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # Define headers with Test Date at the end
    headers = ['Model', 'P/N', 'Serial Number', 'Device Type', 'Processor(s)', 
              'Ram Type', 'Ram Capacity', 'PCIe Devices',
              'Cosmetic Grade', 'Functional Grade', 'Qced by', 'Test Date']
    
    # Style configurations
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                   top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Add headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
        
    # Add data
    ws.cell(row=4, column=1, value=data['model'])
    ws.cell(row=4, column=2, value=data['product_number'])
    ws.cell(row=4, column=3, value=data['serial_number'])
    ws.cell(row=4, column=4, value='Server')
    ws.cell(row=4, column=5, value=data['processor'])
    ws.cell(row=4, column=6, value=data['ram_type'])
    ws.cell(row=4, column=7, value=data['installed_ram'])
    ws.cell(row=4, column=8, value='\n'.join(data['pcie_devices']))
    ws.cell(row=4, column=9, value=data['cosmetic_grade'])
    ws.cell(row=4, column=10, value=data['functional_grade'])
    ws.cell(row=4, column=11, value=data['qced_by'])
    ws.cell(row=4, column=12, value=data['test_date'])
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column[0].column_letter].width = adjusted_width
    
    return wb

def main():
    # Get server information
    print("Gathering server information...")
    server_info = get_server_info()
    
    if not server_info:
        print("Failed to gather server information. Exiting.")
        return
    
    # Get user input
    user_data = get_user_input()
    
    # Combine all data
    all_data = {**server_info, **user_data}
    
    # Create and save spreadsheet
    wb = create_qc_spreadsheet(all_data)
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"Server_QC_Report_{timestamp}.xlsx"
    wb.save(filename)
    print(f"\nQC report saved as: {filename}")
    
    return all_data

if __name__ == "__main__":
    main()
